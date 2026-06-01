#!/usr/bin/env python3
"""Plot latency results from ref_DB/*.xlsx.

The workbook is read directly with openpyxl to avoid requiring pandas.
By default this creates one 2x2 figure with Prefill and Decode batch-size
results as separate subplots.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(__file__).resolve().parent / "ref_DB" / "main_result.xlsx"
DEFAULT_BREAKDOWN_WORKBOOK = Path(__file__).resolve().parent / "ref_DB" / "breakdown_data.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "figures" / "latency_main_result.png"
DEFAULT_BREAKDOWN_OUTPUT = Path(__file__).resolve().parent / "figures" / "breakdown.png"
RESULT_SHEETS = ("Prefill", "Decode_BS1", "Decode_BS8", "Decode_BS64")
BREAKDOWN_SHEETS = ("ISO-AREA", "ISO-POWER")
COMPONENTS = ("qk", "pv", "qkv", "out", "ffn")
CONFIGS = ("C1", "C2", "C3", "C4_alone", "C4_fused")


@dataclass(frozen=True)
class Series:
    label: str
    values: list[float]


@dataclass(frozen=True)
class SheetData:
    title: str
    seq_lengths: list[int]
    series: list[Series]


@dataclass(frozen=True)
class BreakdownSection:
    iso: str
    phase: str
    seq_length: int
    values: dict[str, dict[str, float]]


def parse_seq_length(value: object) -> int:
    """Convert workbook sequence labels like 's512' into integer 512."""
    if value is None:
        raise ValueError("empty sequence value")

    text = str(value).strip().lower()
    if text.startswith("s"):
        text = text[1:]
    return int(text)


def find_header_row(rows: list[tuple[object, ...]]) -> int:
    for row_idx, row in enumerate(rows):
        first_cell = row[0] if row else None
        if isinstance(first_cell, str) and first_cell.strip().lower() == "seq":
            return row_idx
    raise ValueError("could not find header row whose first cell is 'seq'")


def to_float(value: object) -> float:
    if value is None:
        return float("nan")
    return float(value)


def load_sheet_data(workbook: Path, sheet_name: str, include_speedup: bool) -> SheetData:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{workbook} does not contain sheet {sheet_name!r}")

    ws = wb[sheet_name]
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    header_idx = find_header_row(rows)
    header = rows[header_idx]

    seq_lengths: list[int] = []
    data_rows: list[tuple[object, ...]] = []
    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        seq_lengths.append(parse_seq_length(row[0]))
        data_rows.append(row)

    series: list[Series] = []
    for col_idx, name in enumerate(header[1:], start=1):
        if name is None:
            continue

        label = str(name)
        is_speedup = "speedup" in label.lower()
        if is_speedup != include_speedup:
            continue

        values = [to_float(row[col_idx]) if col_idx < len(row) else float("nan") for row in data_rows]
        series.append(Series(label=label, values=values))

    if not series:
        kind = "speedup" if include_speedup else "latency"
        raise ValueError(f"no {kind} series found in sheet {sheet_name!r}")

    title = sheet_name.replace("_", " ")
    return SheetData(title=title, seq_lengths=seq_lengths, series=series)


def parse_breakdown_title(value: object) -> tuple[str, int] | None:
    if not isinstance(value, str) or "·" not in value:
        return None

    phase, seq = [part.strip() for part in value.split("·", maxsplit=1)]
    if phase not in {"Prefill", "Decode"}:
        return None
    return phase, parse_seq_length(seq)


def load_breakdown_sections(workbook: Path) -> list[BreakdownSection]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    sections: list[BreakdownSection] = []

    for sheet_name in BREAKDOWN_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{workbook} does not contain sheet {sheet_name!r}")

        rows = [tuple(row) for row in wb[sheet_name].iter_rows(values_only=True)]
        row_idx = 0
        while row_idx < len(rows):
            title = parse_breakdown_title(rows[row_idx][0] if rows[row_idx] else None)
            if title is None:
                row_idx += 1
                continue

            phase, seq_length = title
            header_idx = row_idx + 1
            if header_idx >= len(rows):
                raise ValueError(f"missing breakdown header after {phase} s{seq_length}")

            header = rows[header_idx]
            config_columns = {
                str(name): col_idx
                for col_idx, name in enumerate(header)
                if isinstance(name, str) and name in CONFIGS
            }
            if set(config_columns) != set(CONFIGS):
                raise ValueError(f"unexpected breakdown header after {phase} s{seq_length}: {header}")

            values = {config: {} for config in CONFIGS}
            row_idx = header_idx + 1
            while row_idx < len(rows):
                row = rows[row_idx]
                component = row[0] if row else None
                if component is None or component == "TOTAL":
                    break
                component_name = str(component)
                if component_name in COMPONENTS:
                    for config, col_idx in config_columns.items():
                        values[config][component_name] = to_float(row[col_idx])
                row_idx += 1

            sections.append(
                BreakdownSection(
                    iso=sheet_name,
                    phase=phase,
                    seq_length=seq_length,
                    values=values,
                )
            )
            row_idx += 1

    return sections


def make_plot(
    sheets: Iterable[SheetData],
    output: Path,
    *,
    title: str,
    ylabel: str,
    xscale: str,
    yscale: str,
) -> None:
    import matplotlib.pyplot as plt

    sheet_list = list(sheets)
    fig, axes = plt.subplots(2, 2, figsize=(15, 9), sharex=True)
    axes_flat = axes.flatten()

    colors = {
        "C1": "#1f77b4",
        "C2": "#ff7f0e",
        "C3": "#2ca02c",
        "C4_alone": "#d62728",
        "C4_fused": "#9467bd",
    }

    for ax, sheet in zip(axes_flat, sheet_list):
        for item in sheet.series:
            config = item.label.split()[0]
            linestyle = "--" if "ISO-POWER" in item.label or "(IP)" in item.label else "-"
            marker = "s" if linestyle == "--" else "o"
            ax.plot(
                sheet.seq_lengths,
                item.values,
                label=item.label,
                color=colors.get(config),
                linestyle=linestyle,
                marker=marker,
                linewidth=1.8,
                markersize=4,
            )

        ax.set_title(sheet.title)
        ax.set_xscale("log", base=2) if xscale == "log2" else ax.set_xscale(xscale)
        ax.set_yscale(yscale)
        ax.set_xticks(sheet.seq_lengths)
        ax.set_xticklabels([str(seq) for seq in sheet.seq_lengths], rotation=35, ha="right")
        ax.set_xlabel("Sequence length")
        ax.set_ylabel(ylabel)
        ax.grid(True, which="both", linestyle=":", linewidth=0.7, alpha=0.65)

    for ax in axes_flat[len(sheet_list) :]:
        ax.axis("off")

    handles, labels = axes_flat[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", ncol=4, frameon=False, bbox_to_anchor=(0.5, 0.99))
    fig.suptitle(title, y=1.05, fontsize=15)
    fig.tight_layout(rect=(0, 0, 1, 0.93))

    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output, dpi=220, bbox_inches="tight")
    plt.close(fig)


def make_breakdown_plot(
    sections: Iterable[BreakdownSection],
    output: Path,
    *,
    value_mode: str,
) -> None:
    import matplotlib.pyplot as plt

    section_map = {(section.iso, section.phase, section.seq_length): section for section in sections}
    ordered_columns = [
        ("Prefill", 512),
        ("Prefill", 16384),
        ("Decode", 512),
        ("Decode", 16384),
    ]
    component_colors = {
        "qk": "#4c78a8",
        "pv": "#f58518",
        "qkv": "#54a24b",
        "out": "#b279a2",
        "ffn": "#e45756",
    }

    fig, axes = plt.subplots(
        len(BREAKDOWN_SHEETS),
        len(ordered_columns),
        figsize=(18, 8),
        sharey=value_mode == "percent",
    )

    for row_idx, iso in enumerate(BREAKDOWN_SHEETS):
        for col_idx, (phase, seq_length) in enumerate(ordered_columns):
            ax = axes[row_idx][col_idx]
            section = section_map.get((iso, phase, seq_length))
            if section is None:
                ax.axis("off")
                continue

            bottoms = [0.0] * len(CONFIGS)
            totals = [
                sum(section.values[config].get(component, 0.0) for component in COMPONENTS)
                for config in CONFIGS
            ]

            for component in COMPONENTS:
                heights = []
                for config_idx, config in enumerate(CONFIGS):
                    raw_value = section.values[config].get(component, 0.0)
                    if value_mode == "percent":
                        total = totals[config_idx]
                        heights.append(raw_value / total * 100.0 if total else 0.0)
                    else:
                        heights.append(raw_value)

                ax.bar(
                    CONFIGS,
                    heights,
                    bottom=bottoms,
                    label=component,
                    color=component_colors[component],
                    edgecolor="white",
                    linewidth=0.6,
                )
                bottoms = [bottom + height for bottom, height in zip(bottoms, heights)]

            ax.set_title(f"{phase} s{seq_length}")
            if col_idx == 0:
                ax.set_ylabel(f"{iso}\n" + ("Component share (%)" if value_mode == "percent" else "Latency (cycles)"))
            ax.grid(True, axis="y", linestyle=":", linewidth=0.7, alpha=0.65)

    handles, labels = axes[0][0].get_legend_handles_labels()
    fig.legend(handles, labels, loc="upper center", ncol=len(COMPONENTS), frameon=False, bbox_to_anchor=(0.5, 1.0))
    suffix = "percent" if value_mode == "percent" else "cycles"
    fig.suptitle(f"Latency Breakdown on HW ({suffix})", y=1.06, fontsize=15)
    fig.tight_layout(rect=(0, 0, 1, 0.93))

    output.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output, dpi=220, bbox_inches="tight")
    plt.close(fig)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Visualize latency data from ref_DB xlsx files.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help=f"input workbook path (default: {DEFAULT_WORKBOOK}; breakdown uses {DEFAULT_BREAKDOWN_WORKBOOK})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=f"output image path (default: {DEFAULT_OUTPUT}; breakdown uses {DEFAULT_BREAKDOWN_OUTPUT})",
    )
    parser.add_argument(
        "--plot",
        choices=("latency", "speedup", "breakdown"),
        default="latency",
        help="plot total latency columns, speedup columns, or component breakdown (default: latency)",
    )
    parser.add_argument(
        "--breakdown-value",
        choices=("cycles", "percent"),
        default="cycles",
        help="breakdown plot value mode (default: cycles)",
    )
    parser.add_argument(
        "--xscale",
        choices=("linear", "log2"),
        default="log2",
        help="x-axis scale for sequence length (default: log2)",
    )
    parser.add_argument(
        "--yscale",
        choices=("linear", "log"),
        default="log",
        help="y-axis scale (default: log)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    default_workbook = DEFAULT_BREAKDOWN_WORKBOOK if args.plot == "breakdown" else DEFAULT_WORKBOOK
    default_output = DEFAULT_BREAKDOWN_OUTPUT if args.plot == "breakdown" else DEFAULT_OUTPUT
    args.workbook = default_workbook if args.workbook is None else args.workbook
    args.output = default_output if args.output is None else args.output

    workbook = args.workbook.resolve()
    if not workbook.exists():
        raise FileNotFoundError(workbook)

    if args.plot == "breakdown":
        sections = load_breakdown_sections(workbook)
        make_breakdown_plot(sections, args.output, value_mode=args.breakdown_value)
        print(f"wrote {args.output}")
        return

    include_speedup = args.plot == "speedup"
    sheets = [load_sheet_data(workbook, sheet_name, include_speedup) for sheet_name in RESULT_SHEETS]
    title = "Latency on HW" if args.plot == "latency" else "Speedup on HW"
    ylabel = "Latency (cycles)" if args.plot == "latency" else "Speedup vs C1"

    make_plot(
        sheets,
        args.output,
        title=title,
        ylabel=ylabel,
        xscale=args.xscale,
        yscale=args.yscale,
    )
    print(f"wrote {args.output}")


if __name__ == "__main__":
    main()
